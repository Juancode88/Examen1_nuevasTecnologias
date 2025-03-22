import openpyxl

# Parte 1: Crear un archivo Excel básico
def crear_archivo_excel():
    # Crear un archivo nuevo
    libro = openpyxl.Workbook()
    hoja = libro.active

    # Escribir en celdas
    hoja["A1"] = "Hola"
    hoja["B1"] = "Mundo"

    # Guardar el archivo
    libro.save("mi_primer_excel.xlsx")
    print("¡Archivo 'mi_primer_excel.xlsx' creado!")

# Parte 2: Leer y modificar el archivo Excel
def leer_y_modificar_excel():
    # Abrir el archivo
    libro = openpyxl.load_workbook("mi_primer_excel.xlsx")
    hoja = libro.active

    # Leer lo que está en A1 y B1
    print(f"Valor en A1: {hoja['A1'].value}")
    print(f"Valor en B1: {hoja['B1'].value}")

    # Cambiar lo que está en B1
    hoja["B1"] = "Amigos"

    # Guardar el cambio
    libro.save("mi_primer_excel.xlsx")
    print("¡Cambio en 'mi_primer_excel.xlsx' hecho!")

# Parte 3: Ejercicio de estudiantes aprobados
def estudiantes_aprobados():
    # Crear diccionario y entrada de datos
    estudiantes = {}
    for i in range(3):
        nombre = input(f"Ingrese el nombre del estudiante {i+1}: ")
        nota = float(input(f"Ingrese la nota del estudiante {i+1}: "))
        estudiantes[nombre] = nota

    # Crear archivo Excel
    libro = openpyxl.Workbook()
    hoja = libro.active

    # Escribir encabezado
    hoja["A1"] = "Aprobados (>=60)"

    # Escribir aprobados con ciclo y condicional
    fila = 2
    for nombre, nota in estudiantes.items():
        if nota >= 60:
            hoja[f"A{fila}"] = nombre
            fila += 1

    # Guardar archivo
    libro.save("ejercicio2.xlsx")
    print("¡Ejercicio 2 guardado en 'ejercicio2.xlsx'!")

# Ejecutar todas las partes
crear_archivo_excel()
leer_y_modificar_excel()
estudiantes_aprobados()